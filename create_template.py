from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

wb = Workbook()

# 样式
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
header_font = Font(bold=True, size=11, color="FFFFFF")
thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

# ===== 门店信息模板 =====
ws1 = wb.active
ws1.title = "门店信息"

headers1 = ["门店名称", "可容纳人数", "适配范围"]
for col, h in enumerate(headers1, 1):
    cell = ws1.cell(row=1, column=col, value=h)
    cell.font = header_font; cell.fill = header_fill; cell.alignment = Alignment(horizontal='center'); cell.border = thin_border

examples1 = [
    ["小米之家北京海淀区大悦城专卖店", 3, "服务"],
    ["小米之家北京朝阳区三里屯太古里店", 2, "非服务"],
]
for r, row in enumerate(examples1, 2):
    for c, v in enumerate(row, 1):
        cell = ws1.cell(row=r, column=c, value=v); cell.border = thin_border

ws1.cell(row=5, column=1, value="说明：").font = Font(bold=True, color="FF0000")
ws1.cell(row=6, column=1, value="门店名称：必填")
ws1.cell(row=7, column=1, value="可容纳人数：必填，数字")
ws1.cell(row=8, column=1, value='适配范围：必填，"服务"或"非服务"')
ws1.cell(row=9, column=1, value="请删除示例数据后填写")
ws1.column_dimensions['A'].width = 45; ws1.column_dimensions['B'].width = 15; ws1.column_dimensions['C'].width = 15

# ===== 学员信息模板 =====
ws2 = wb.create_sheet("学员信息")

headers2 = ["工号", "姓名", "类型", "职位名称", "一级部门", "二级部门", "三级部门", "四级部门", "五级部门"]
for col, h in enumerate(headers2, 1):
    cell = ws2.cell(row=1, column=col, value=h)
    cell.font = header_font; cell.fill = header_fill; cell.alignment = Alignment(horizontal='center'); cell.border = thin_border

examples2 = [
    ["S001", "张三", "服务", "服务工程师", "中国区", "服务部", "北京服务中心", "海淀服务站", ""],
    ["S002", "李四", "服务", "店员", "中国区", "服务部", "北京服务中心", "朝阳服务站", ""],
    ["N001", "王五", "非服务", "运营专员", "中国区", "运营部", "北京运营中心", "", ""],
    ["N002", "赵六", "非服务", "销售顾问", "中国区", "销售部", "北京销售中心", "海淀销售站", ""],
]
for r, row in enumerate(examples2, 2):
    for c, v in enumerate(row, 1):
        cell = ws2.cell(row=r, column=c, value=v); cell.border = thin_border

ws2.cell(row=8, column=1, value="说明：").font = Font(bold=True, color="FF0000")
ws2.cell(row=9, column=1, value="工号、姓名、类型：必填")
ws2.cell(row=10, column=1, value="职位名称、一级~五级部门：选填")
ws2.cell(row=11, column=1, value='类型：填"服务"或"非服务"')
ws2.cell(row=12, column=1, value="请删除示例数据后填写")

ws2.column_dimensions['A'].width = 12; ws2.column_dimensions['B'].width = 10; ws2.column_dimensions['C'].width = 10
ws2.column_dimensions['D'].width = 15; ws2.column_dimensions['E'].width = 12; ws2.column_dimensions['F'].width = 12
ws2.column_dimensions['G'].width = 15; ws2.column_dimensions['H'].width = 15; ws2.column_dimensions['I'].width = 15

wb.save(r"C:\Users\Sunny Song\大聪明和Sunny的AI探索\projects\store-allocation\导入模板.xlsx")
print("模板已生成")
